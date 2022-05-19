import openpyxl


def getRowCount(file, sheetName):
    wb = openpyxl.load_workbook(file)
    ws = wb[sheetName]
    return ws.max_row


def readData(file, sheetName, rowNo, colNo):
    wb = openpyxl.load_workbook(file)
    ws = wb[sheetName]
    return ws.cell(row=rowNo, column=colNo).value
